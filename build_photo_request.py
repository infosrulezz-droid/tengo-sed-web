#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_photo_request.py — genera FOTOS_QUE_NECESITO.xlsx

La lista de fotos de producto que hay que conseguir, ordenada por urgencia.
Se regenera solo, asi que despues de arreglar fotos volves a correrlo y la
lista se achica.

    python build_photo_request.py
"""
import json, io, os
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))

# Verificado a mano abriendo cada imagen — producto equivocado, no solo fea.
WRONG = {
    "120 Spritz tinto 1.5 lt":        "Muestra APERITIVO ROSATO. Los 4 archivos de 120 Spritz son identicos.",
    "120 Spritz Ponche 1.5 lt":       "Muestra Rosato, no Ponche.",
    "120 Spritz Biotropical 1.5 lt":  "Muestra Rosato, no Biotropical.",
    "Gato Dulce Chocolate 1.5 lt":    "Identica a Chirimoya, y la etiqueta dice Torontel. No es ninguno de los dos.",
    "Sprite 3 Lts":                   "La foto es de una botella de 1.5 L.",
    "MAS CACHANTUN SABORES 1.5LT":    "Duplica la foto de Mas Citrus.",
    "Ron Barcelo Dorado 1lt":         "Muestra Barcelo AMAIA, otra expresion.",
    "Ron Barceló Dorado 1lt":         "Muestra Barcelo AMAIA, otra expresion.",
    "Royal Guard Bot 650 cc":         "La etiqueta de la foto dice 355 cc.",
    "Cachantun c/g 1.5lt":            "La etiqueta dice 1.6 L.",
    "Austral Patagonia Mix Whisky 470cc": "La etiqueta dice lata 350 cc.",
    "Pisco Alto del Carmen 1LT":      "La foto es del formato 750 ML.",
    "Alto Carmen 35 1lt":             "Comparte foto con Pisco Alto del Carmen; ademas parece entrada duplicada.",
}

PRIO = {"1 - Producto equivocado": "C00000",
        "2 - Muy baja resolucion": "ED7D31",
        "3 - Baja resolucion":     "FFC000"}


def main():
    m = json.load(io.open(os.path.join(ROOT, "product_img_map.json"), encoding="utf-8"))
    d = json.load(io.open(os.path.join(ROOT, "products_inventory.json"), encoding="utf-8"))
    items = d if isinstance(d, list) else d.get("products", d)

    rows = []
    for p in items:
        name = p.get("name", "")
        v = m.get(name)
        if not v:
            continue
        fn = v.split("?")[0]
        path = os.path.join(ROOT, "products", fn)
        if not os.path.exists(path):
            continue
        try:
            with Image.open(path) as im:
                w, h = im.size
        except Exception:
            w = h = 0
        edge = max(w, h)

        if name in WRONG:
            prio, why = "1 - Producto equivocado", WRONG[name]
        elif edge < 200:
            prio, why = "2 - Muy baja resolucion", f"Solo {w}x{h} px. Se ve borrosa en la grilla."
        elif edge < 250:
            prio, why = "3 - Baja resolucion", f"{w}x{h} px. Justa; se nota en pantallas retina."
        else:
            continue
        rows.append([prio, name, p.get("cat", ""), p.get("price", ""), fn, f"{w}x{h}", why])

    rows.sort(key=lambda r: (r[0], -0 if r[0].startswith("1") else 0, r[1]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Fotos que necesito"

    ws["A1"] = "Tengo Sed — fotos de producto por conseguir"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = ("Lo que se necesita es la foto OFICIAL de la marca o distribuidor, "
                "en la mayor resolucion posible. Minimo 600x600 px.")
    ws["A2"].font = Font(size=10, italic=True, color="555555")
    ws["A3"] = ("Verificar marca, formato (1.5 lt no es 750 cc) y variedad/sabor. "
                "Una foto linda del producto equivocado no sirve.")
    ws["A3"].font = Font(size=10, italic=True, color="555555")

    head = ["Prioridad", "Producto", "Categoria", "Precio CLP",
            "Archivo actual", "Tamano actual", "Que pasa"]
    hrow = 5
    for i, h in enumerate(head, 1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row in enumerate(rows, hrow + 1):
        for i, val in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(i == 7))
            if i == 1:
                c.font = Font(bold=True, color=PRIO.get(val, "000000"))
            if i == 4 and isinstance(val, (int, float)):
                c.number_format = '#,##0'

    for col, wdt in zip("ABCDEFG", (22, 34, 12, 12, 40, 14, 52)):
        ws.column_dimensions[col].width = wdt
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    ws.auto_filter.ref = f"A{hrow}:G{hrow+len(rows)}"

    # ---- resumen ----
    ws2 = wb.create_sheet("Resumen")
    counts = {}
    for r in rows:
        counts[r[0]] = counts.get(r[0], 0) + 1
    ws2["A1"] = "Resumen"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2.append([])
    ws2.append(["Prioridad", "Fotos"])
    for c in ws2[3]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
    for k in sorted(counts):
        ws2.append([k, counts[k]])
    ws2.append([])
    ws2.append(["TOTAL", len(rows)])
    ws2[f"A{ws2.max_row}"].font = Font(bold=True)
    ws2[f"B{ws2.max_row}"].font = Font(bold=True)
    ws2.append([])
    ws2.append(["Nota", "Prioridad 1 son productos donde la foto muestra OTRO producto."])
    ws2.append(["", "Eso confunde al cliente y puede generar un reclamo. Van primero."])
    ws2.append(["", "Prioridad 2 y 3 son fotos correctas pero chicas: se ven borrosas."])
    ws2.append(["", "No se arreglan agrandandolas — hace falta la foto original."])
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 68

    out = os.path.join(ROOT, "FOTOS_QUE_NECESITO.xlsx")
    wb.save(out)
    print(f"{out}")
    for k in sorted(counts):
        print(f"  {k}: {counts[k]}")
    print(f"  TOTAL: {len(rows)}")


if __name__ == "__main__":
    main()
